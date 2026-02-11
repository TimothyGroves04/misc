#!/usr/bin/env python3
"""
Generate a 3-way financial model for Transurban Group (ASX: TCL).

Creates an Excel workbook with:
  - Assumptions & Drivers
  - Income Statement
  - Balance Sheet
  - Cash Flow Statement

Historical data covers FY21-FY25 (years ended 30 June).
Forecasts cover FY26-FY30 and are formula-driven from the Assumptions tab.
All three statements link to each other following best-practice modelling.

Figures are in Australian Dollars, millions (A$m).
"""

import os
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 0.  Workbook & style helpers
# ---------------------------------------------------------------------------

wb = Workbook()

# Colour palette
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
FORECAST_BG = "FFF2CC"  # light yellow for forecast columns
HIST_BG = "D6E4F0"  # light blue for historical columns
GREEN_FONT = "006100"
DARK_FONT = "1F3864"

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
section_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
input_font = Font(name="Calibri", size=11, color=MED_BLUE)
total_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
pct_fmt = '0.0%'
num_fmt = '#,##0'
num_fmt_1dp = '#,##0.0'
acct_fmt = '#,##0;(#,##0);"-"'

thin_border = Border(
    bottom=Side(style="thin", color="B4C6E7"),
)
thick_bottom = Border(
    bottom=Side(style="medium", color=DARK_BLUE),
)
double_bottom = Border(
    bottom=Side(style="double", color=DARK_BLUE),
)

forecast_fill = PatternFill(start_color=FORECAST_BG, end_color=FORECAST_BG, fill_type="solid")


def style_header_row(ws, row, max_col):
    """Apply dark-blue header style to a row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_row(ws, row, max_col):
    """Apply section header style."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill


def apply_number_format(ws, row, col_start, col_end, fmt):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).number_format = fmt


def mark_forecast_cols(ws, row, col_start, col_end):
    """Lightly shade forecast columns."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if cell.fill == PatternFill() or cell.fill.start_color.index == "00000000":
            cell.fill = forecast_fill


def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ---------------------------------------------------------------------------
# Column layout: A = labels, B = units, C-G = FY21-FY25 (hist), H-L = FY26-FY30
# ---------------------------------------------------------------------------
LABEL_COL = 1  # A
UNIT_COL = 2   # B
HIST_START = 3  # C  = FY21
HIST_END = 7    # G  = FY25
FC_START = 8    # H  = FY26
FC_END = 12     # L  = FY30
MAX_COL = FC_END

FY_LABELS = ["FY21", "FY22", "FY23", "FY24", "FY25",
             "FY26F", "FY27F", "FY28F", "FY29F", "FY30F"]

# ---------------------------------------------------------------------------
# 1.  ASSUMPTIONS & DRIVERS  (Sheet 1)
# ---------------------------------------------------------------------------
ws_a = wb.active
ws_a.title = "Assumptions"
ws_a.sheet_properties.tabColor = MED_BLUE

# Title
ws_a.merge_cells("A1:L1")
ws_a["A1"] = "Transurban Group (ASX: TCL) – Forecast Assumptions & Key Drivers"
ws_a["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws_a["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_a.row_dimensions[1].height = 30

ws_a.merge_cells("A2:L2")
ws_a["A2"] = "All figures in A$ millions unless otherwise stated.  Fiscal year ends 30 June."
ws_a["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

# -- Row 4: SCENARIO TOGGLES section header --
ws_a.cell(row=4, column=LABEL_COL, value="SCENARIO TOGGLES")
style_section_row(ws_a, 4, MAX_COL)

toggle_font = Font(name="Calibri", size=11, color=GREEN_FONT)
toggle_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

toggle_items = [
    (5, "Toll revenue growth method"),
    (6, "Other revenue growth method"),
    (7, "Opex % method"),
    (8, "D&A % method"),
    (9, "Cost of debt method"),
    (10, "Tax rate method"),
]
for trow, tlabel in toggle_items:
    ws_a.cell(row=trow, column=LABEL_COL, value=tlabel).font = normal_font
    cell_b = ws_a.cell(row=trow, column=UNIT_COL, value="Forecast")
    cell_b.font = toggle_font
    cell_b.fill = toggle_fill

# Row 12: Column headers
r = 12
ws_a.cell(row=r, column=LABEL_COL, value="Assumption / Driver")
ws_a.cell(row=r, column=UNIT_COL, value="Unit")
for i, fy in enumerate(FY_LABELS):
    ws_a.cell(row=r, column=HIST_START + i, value=fy)
style_header_row(ws_a, r, MAX_COL)


def _write_values(ws, row, vals, is_pct, fmt=None):
    """Write a list of 10 values to columns C-L."""
    for i, v in enumerate(vals):
        col = HIST_START + i
        cell = ws.cell(row=row, column=col)
        if v is not None:
            if is_pct:
                cell.value = v / 100.0
                cell.number_format = pct_fmt
            else:
                cell.value = v
                cell.number_format = fmt if fmt else num_fmt
            cell.font = input_font if col >= FC_START else normal_font
            if col >= FC_START:
                cell.fill = forecast_fill


def _write_active_row(ws, active_row, input_row, toggle_cell):
    """Write an active row: historical cols reference input, forecast cols use IF."""
    for i in range(5):  # historical columns
        col = HIST_START + i
        cl = get_column_letter(col)
        cell = ws.cell(row=active_row, column=col)
        cell.value = f"={cl}{input_row}"
        cell.number_format = pct_fmt
        cell.font = normal_font
    # Last 3 historical columns for 3yr average
    avg_start = get_column_letter(HIST_END - 2)
    avg_end = get_column_letter(HIST_END)
    for i in range(5):  # forecast columns
        col = FC_START + i
        cl = get_column_letter(col)
        cell = ws.cell(row=active_row, column=col)
        cell.value = f'=IF({toggle_cell}="3yr Avg",AVERAGE({avg_start}{input_row}:{avg_end}{input_row}),{cl}{input_row})'
        cell.number_format = pct_fmt
        cell.font = input_font
        cell.fill = forecast_fill


# Row 14: REVENUE DRIVERS
ws_a.cell(row=14, column=LABEL_COL, value="REVENUE DRIVERS")
style_section_row(ws_a, 14, MAX_COL)

# Row 15: Toll revenue growth (input)
ws_a.cell(row=15, column=LABEL_COL, value="Toll revenue growth (input)")
ws_a.cell(row=15, column=UNIT_COL, value="%")
_write_values(ws_a, 15, [-4.2, 15.1, 22.1, 8.7, 6.2, 5.5, 5.0, 4.5, 4.0, 3.8], True)

# Row 16: Toll revenue growth (active)
ws_a.cell(row=16, column=LABEL_COL, value="Toll revenue growth (active)")
ws_a.cell(row=16, column=UNIT_COL, value="%")
_write_active_row(ws_a, 16, 15, "$B$5")

# Row 17: Other revenue growth (input)
ws_a.cell(row=17, column=LABEL_COL, value="Other revenue growth (input)")
ws_a.cell(row=17, column=UNIT_COL, value="%")
_write_values(ws_a, 17, [2.0, -2.5, 6.8, -1.8, 10.4, 3.0, 3.0, 3.0, 3.0, 3.0], True)

# Row 18: Other revenue growth (active)
ws_a.cell(row=18, column=LABEL_COL, value="Other revenue growth (active)")
ws_a.cell(row=18, column=UNIT_COL, value="%")
_write_active_row(ws_a, 18, 17, "$B$6")

# Row 20: OPERATING COST DRIVERS
ws_a.cell(row=20, column=LABEL_COL, value="OPERATING COST DRIVERS")
style_section_row(ws_a, 20, MAX_COL)

# Row 21: Opex as % of revenue (input)
ws_a.cell(row=21, column=LABEL_COL, value="Opex as % of revenue (input)")
ws_a.cell(row=21, column=UNIT_COL, value="%")
_write_values(ws_a, 21, [35.2, 32.5, 31.5, 30.7, 30.8, 30.5, 30.0, 29.5, 29.0, 28.8], True)

# Row 22: Opex as % of revenue (active)
ws_a.cell(row=22, column=LABEL_COL, value="Opex as % of revenue (active)")
ws_a.cell(row=22, column=UNIT_COL, value="%")
_write_active_row(ws_a, 22, 21, "$B$7")

# Row 23: Employee costs as % of revenue
ws_a.cell(row=23, column=LABEL_COL, value="Employee costs as % of revenue")
ws_a.cell(row=23, column=UNIT_COL, value="%")
_write_values(ws_a, 23, [8.5, 8.0, 7.4, 7.3, 7.2, 7.2, 7.0, 6.9, 6.8, 6.7], True)

# Row 24: Employee costs share of total opex
ws_a.cell(row=24, column=LABEL_COL, value="Employee costs share of total opex")
ws_a.cell(row=24, column=UNIT_COL, value="%")
_write_values(ws_a, 24, [24, 25, 24, 24, 23, 25, 25, 25, 25, 25], True)

# Row 25: Road ops share of total opex
ws_a.cell(row=25, column=LABEL_COL, value="Road ops share of total opex")
ws_a.cell(row=25, column=UNIT_COL, value="%")
_write_values(ws_a, 25, [34, 33, 33, 32, 31, 33, 33, 33, 33, 33], True)

# Row 27: DEPRECIATION & AMORTISATION
ws_a.cell(row=27, column=LABEL_COL, value="DEPRECIATION & AMORTISATION")
style_section_row(ws_a, 27, MAX_COL)

# Row 28: D&A as % of NCA (input)
ws_a.cell(row=28, column=LABEL_COL, value="D&A as % of NCA (input)")
ws_a.cell(row=28, column=UNIT_COL, value="%")
_write_values(ws_a, 28, [2.7, 2.7, 2.7, 2.7, 2.8, 2.8, 2.8, 2.8, 2.8, 2.8], True)

# Row 29: D&A as % of NCA (active)
ws_a.cell(row=29, column=LABEL_COL, value="D&A as % of NCA (active)")
ws_a.cell(row=29, column=UNIT_COL, value="%")
_write_active_row(ws_a, 29, 28, "$B$8")

# Row 30: D&A allocation to PP&E
ws_a.cell(row=30, column=LABEL_COL, value="D&A allocation to PP&E")
ws_a.cell(row=30, column=UNIT_COL, value="%")
_write_values(ws_a, 30, [40, 40, 40, 40, 40, 40, 40, 40, 40, 40], True)

# Row 31: D&A allocation to intangibles
ws_a.cell(row=31, column=LABEL_COL, value="D&A allocation to intangibles")
ws_a.cell(row=31, column=UNIT_COL, value="%")
_write_values(ws_a, 31, [60, 60, 60, 60, 60, 60, 60, 60, 60, 60], True)

# Row 33: FINANCING ASSUMPTIONS
ws_a.cell(row=33, column=LABEL_COL, value="FINANCING ASSUMPTIONS")
style_section_row(ws_a, 33, MAX_COL)

# Row 34: Average cost of debt (input)
ws_a.cell(row=34, column=LABEL_COL, value="Average cost of debt (input)")
ws_a.cell(row=34, column=UNIT_COL, value="%")
_write_values(ws_a, 34, [4.6, 4.0, 4.2, 4.5, 4.6, 4.7, 4.7, 4.8, 4.8, 4.8], True)

# Row 35: Average cost of debt (active)
ws_a.cell(row=35, column=LABEL_COL, value="Average cost of debt (active)")
ws_a.cell(row=35, column=UNIT_COL, value="%")
_write_active_row(ws_a, 35, 34, "$B$9")

# Row 36: Effective tax rate (input)
ws_a.cell(row=36, column=LABEL_COL, value="Effective tax rate (input)")
ws_a.cell(row=36, column=UNIT_COL, value="%")
_write_values(ws_a, 36, [-1.5, 14.0, 14.3, 14.4, 15.0, 15.0, 15.0, 15.0, 15.0, 15.0], True)

# Row 37: Effective tax rate (active)
ws_a.cell(row=37, column=LABEL_COL, value="Effective tax rate (active)")
ws_a.cell(row=37, column=UNIT_COL, value="%")
_write_active_row(ws_a, 37, 36, "$B$10")

# Row 38: Statutory tax rate
ws_a.cell(row=38, column=LABEL_COL, value="Statutory tax rate")
ws_a.cell(row=38, column=UNIT_COL, value="%")
_write_values(ws_a, 38, [30, 30, 30, 30, 30, 30, 30, 30, 30, 30], True)

# Row 39: Current borrowings as % of total debt
ws_a.cell(row=39, column=LABEL_COL, value="Current borrowings as % of total debt")
ws_a.cell(row=39, column=UNIT_COL, value="%")
_write_values(ws_a, 39, [6, 5.5, 5.5, 5, 5, 5, 5, 5, 5, 5], True)

# Row 40: DRP / dilution rate
ws_a.cell(row=40, column=LABEL_COL, value="DRP / dilution rate")
ws_a.cell(row=40, column=UNIT_COL, value="%")
_write_values(ws_a, 40, [1, 1, 1, 1, 1, 1, 1, 1, 1, 1], True)

# Row 42: BALANCE SHEET DRIVERS
ws_a.cell(row=42, column=LABEL_COL, value="BALANCE SHEET DRIVERS")
style_section_row(ws_a, 42, MAX_COL)

# Row 43: Capex
ws_a.cell(row=43, column=LABEL_COL, value="Capex (maintenance + growth)")
ws_a.cell(row=43, column=UNIT_COL, value="A$m")
_write_values(ws_a, 43, [628, 1_092, 1_805, 1_420, 1_200, 1_300, 1_350, 1_400, 1_250, 1_200], False)

# Row 44: Trade receivables days
ws_a.cell(row=44, column=LABEL_COL, value="Trade receivables days")
ws_a.cell(row=44, column=UNIT_COL, value="days")
_write_values(ws_a, 44, [28, 26, 25, 24, 24, 24, 24, 24, 24, 24], False)

# Row 45: Trade payables days
ws_a.cell(row=45, column=LABEL_COL, value="Trade payables days")
ws_a.cell(row=45, column=UNIT_COL, value="days")
_write_values(ws_a, 45, [55, 52, 50, 48, 48, 48, 48, 48, 48, 48], False)

# Row 46: Other current assets growth rate
ws_a.cell(row=46, column=LABEL_COL, value="Other current assets growth rate")
ws_a.cell(row=46, column=UNIT_COL, value="%")
_write_values(ws_a, 46, [3, 3, 3, 3, 3, 3, 3, 3, 3, 3], True)

# Row 47: JV investments growth rate
ws_a.cell(row=47, column=LABEL_COL, value="JV investments growth rate")
ws_a.cell(row=47, column=UNIT_COL, value="%")
_write_values(ws_a, 47, [-2, -2, -2, -2, -2, -2, -2, -2, -2, -2], True)

# Row 48: Other NCA growth rate
ws_a.cell(row=48, column=LABEL_COL, value="Other NCA growth rate")
ws_a.cell(row=48, column=UNIT_COL, value="%")
_write_values(ws_a, 48, [2, 2, 2, 2, 2, 2, 2, 2, 2, 2], True)

# Row 49: Other current liabilities growth rate
ws_a.cell(row=49, column=LABEL_COL, value="Other current liabilities growth rate")
ws_a.cell(row=49, column=UNIT_COL, value="%")
_write_values(ws_a, 49, [3, 3, 3, 3, 3, 3, 3, 3, 3, 3], True)

# Row 50: Other NCL growth rate
ws_a.cell(row=50, column=LABEL_COL, value="Other NCL growth rate")
ws_a.cell(row=50, column=UNIT_COL, value="%")
_write_values(ws_a, 50, [2, 2, 2, 2, 2, 2, 2, 2, 2, 2], True)

# Row 52: CASH FLOW / CAPITAL STRUCTURE
ws_a.cell(row=52, column=LABEL_COL, value="CASH FLOW / CAPITAL STRUCTURE")
style_section_row(ws_a, 52, MAX_COL)

# Row 53: DPS
ws_a.cell(row=53, column=LABEL_COL, value="Dividend per security (DPS)")
ws_a.cell(row=53, column=UNIT_COL, value="A¢")
_write_values(ws_a, 53, [41.0, 53.0, 62.0, 64.5, 66.0, 68.0, 70.0, 72.0, 74.0, 76.0], False, num_fmt_1dp)

# Row 54: Securities on issue
ws_a.cell(row=54, column=LABEL_COL, value="Securities on issue (approx)")
ws_a.cell(row=54, column=UNIT_COL, value="m")
_write_values(ws_a, 54, [1_932, 1_948, 1_964, 1_978, 1_990, 2_000, 2_010, 2_020, 2_025, 2_030], False)

# Row 55: Net debt issuance
ws_a.cell(row=55, column=LABEL_COL, value="Net debt issuance / (repayment)")
ws_a.cell(row=55, column=UNIT_COL, value="A$m")
_write_values(ws_a, 55, [1_200, 2_050, 2_800, 600, 300, 400, 350, 300, 200, 100], False)

# Row 56: Other operating adj. growth rate
ws_a.cell(row=56, column=LABEL_COL, value="Other operating adj. growth rate")
ws_a.cell(row=56, column=UNIT_COL, value="%")
_write_values(ws_a, 56, [2, 2, 2, 2, 2, 2, 2, 2, 2, 2], True)

# Row 57: Equity issuance rate (DRP)
ws_a.cell(row=57, column=LABEL_COL, value="Equity issuance rate (DRP)")
ws_a.cell(row=57, column=UNIT_COL, value="%")
_write_values(ws_a, 57, [1, 1, 1, 1, 1, 1, 1, 1, 1, 1], True)

# Row 59: NOTES ASSUMPTIONS
ws_a.cell(row=59, column=LABEL_COL, value="NOTES ASSUMPTIONS")
style_section_row(ws_a, 59, MAX_COL)

# Row 60: Construction revenue growth
ws_a.cell(row=60, column=LABEL_COL, value="Construction revenue growth")
ws_a.cell(row=60, column=UNIT_COL, value="%")
_write_values(ws_a, 60, [3, 3, 3, 3, 3, 3, 3, 3, 3, 3], True)

# Row 61: Debt maturity 1-2yr as % of total
ws_a.cell(row=61, column=LABEL_COL, value="Debt maturity 1-2yr as % of total")
ws_a.cell(row=61, column=UNIT_COL, value="%")
_write_values(ws_a, 61, [5.3, 4.9, 4.8, 4.7, 4.2, 5, 5, 5, 5, 5], True)

# Row 62: Debt maturity 2-5yr as % of total
ws_a.cell(row=62, column=LABEL_COL, value="Debt maturity 2-5yr as % of total")
ws_a.cell(row=62, column=UNIT_COL, value="%")
_write_values(ws_a, 62, [25, 22, 22, 26, 26, 26, 26, 26, 26, 26], True)

# Row 63: Capitalised borrowing costs factor
ws_a.cell(row=63, column=LABEL_COL, value="Capitalised borrowing costs factor")
ws_a.cell(row=63, column=UNIT_COL, value="%")
_write_values(ws_a, 63, [15, 15, 15, 15, 15, 15, 15, 15, 15, 15], True)

# Row 64: Forecast non-deductible amortisation
ws_a.cell(row=64, column=LABEL_COL, value="Forecast non-deductible amortisation")
ws_a.cell(row=64, column=UNIT_COL, value="A$m")
_write_values(ws_a, 64, [190, 190, 190, 190, 190, 190, 190, 190, 190, 190], False)

# Row 65: Forecast tax concessions
ws_a.cell(row=65, column=LABEL_COL, value="Forecast tax concessions")
ws_a.cell(row=65, column=UNIT_COL, value="A$m")
_write_values(ws_a, 65, [-40, -40, -40, -40, -40, -40, -40, -40, -40, -40], False)

# Row 66: Forecast other perm. differences
ws_a.cell(row=66, column=LABEL_COL, value="Forecast other perm. differences")
ws_a.cell(row=66, column=UNIT_COL, value="A$m")
_write_values(ws_a, 66, [8, 8, 8, 8, 8, 8, 8, 8, 8, 8], False)

# Row 67: Capital commitments multiple of capex
ws_a.cell(row=67, column=LABEL_COL, value="Capital commitments multiple of capex")
ws_a.cell(row=67, column=UNIT_COL, value="x")
_write_values(ws_a, 67, [1.2, 1.2, 1.2, 1.2, 1.2, 1.2, 1.2, 1.2, 1.2, 1.2], False, num_fmt_1dp)

# Row 68: Operating lease growth rate
ws_a.cell(row=68, column=LABEL_COL, value="Operating lease growth rate")
ws_a.cell(row=68, column=UNIT_COL, value="%")
_write_values(ws_a, 68, [3, 3, 3, 3, 3, 3, 3, 3, 3, 3], True)

# Row 69: Forecast contingent liabilities
ws_a.cell(row=69, column=LABEL_COL, value="Forecast contingent liabilities")
ws_a.cell(row=69, column=UNIT_COL, value="A$m")
_write_values(ws_a, 69, [170, 170, 170, 170, 170, 170, 170, 170, 170, 170], False)

# Store assumption row constants for referencing from other sheets
AROW_TOLL_GR = 16        # Toll revenue growth (active)
AROW_OTHER_GR = 18       # Other revenue growth (active)
AROW_OPEX_PCT = 22       # Opex % (active)
AROW_EMP_PCT = 23        # Employee costs % of revenue
AROW_EMP_SHARE = 24      # Employee costs share of total opex
AROW_ROAD_SHARE = 25     # Road ops share of total opex
AROW_DA_PCT = 29         # D&A % of NCA (active)
AROW_DA_PPE = 30         # D&A allocation to PP&E
AROW_DA_INTANG = 31      # D&A allocation to intangibles
AROW_COD = 35            # Avg cost of debt (active)
AROW_TAX = 37            # Effective tax rate (active)
AROW_STAT_TAX = 38       # Statutory tax rate
AROW_CURR_BORROW_PCT = 39  # Current borrowings % of total debt
AROW_DRP_RATE = 40       # DRP / dilution rate
AROW_CAPEX = 43          # Capex
AROW_REC_DAYS = 44       # Trade receivables days
AROW_PAY_DAYS = 45       # Trade payables days
AROW_OCA_GR = 46         # Other current assets growth
AROW_JV_GR = 47          # JV investments growth
AROW_ONCA_GR = 48        # Other NCA growth
AROW_OCL_GR = 49         # Other current liabilities growth
AROW_ONCL_GR = 50        # Other NCL growth
AROW_DPS = 53            # DPS
AROW_SHARES = 54         # Securities on issue
AROW_NET_DEBT = 55       # Net debt issuance
AROW_OTHER_OPS_GR = 56   # Other operating adj. growth
AROW_EQUITY_ISS = 57     # Equity issuance rate (DRP)
AROW_CONSTR_GR = 60      # Construction revenue growth
AROW_MAT_1_2 = 61        # Debt maturity 1-2yr %
AROW_MAT_2_5 = 62        # Debt maturity 2-5yr %
AROW_CAP_BORR = 63       # Capitalised borrowing costs factor
AROW_NON_DED = 64        # Non-deductible amortisation
AROW_TAX_CONC = 65       # Tax concessions
AROW_OTHER_PERM = 66     # Other perm differences
AROW_CAP_COMMIT = 67     # Capital commitments multiple
AROW_OP_LEASE_GR = 68    # Operating lease growth
AROW_CONTINGENT = 69     # Contingent liabilities

set_col_widths(ws_a, {"A": 40, "B": 10})
for i in range(HIST_START, MAX_COL + 1):
    ws_a.column_dimensions[get_column_letter(i)].width = 14

# ---------------------------------------------------------------------------
# 2.  INCOME STATEMENT  (Sheet 2)
# ---------------------------------------------------------------------------
ws_is = wb.create_sheet("Income Statement")
ws_is.sheet_properties.tabColor = "4472C4"

ws_is.merge_cells("A1:L1")
ws_is["A1"] = "Transurban Group – Income Statement"
ws_is["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws_is.row_dimensions[1].height = 30

ws_is.merge_cells("A2:L2")
ws_is["A2"] = "A$ millions  |  Fiscal year ends 30 June"
ws_is["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

# Header row (row 4)
r = 4
ws_is.cell(row=r, column=LABEL_COL, value="Income Statement")
ws_is.cell(row=r, column=UNIT_COL, value="")
for i, fy in enumerate(FY_LABELS):
    ws_is.cell(row=r, column=HIST_START + i, value=fy)
style_header_row(ws_is, r, MAX_COL)

# Historical data  (FY21-FY25)
hist_toll_rev = [2_459, 2_830, 3_455, 3_756, 3_990]
hist_other_rev = [319, 311, 332, 326, 360]
hist_total_rev = [t + o for t, o in zip(hist_toll_rev, hist_other_rev)]  # derived

hist_employee = [-236, -251, -280, -298, -313]
hist_road_ops = [-340, -333, -380, -395, -418]
hist_corp_admin = [-401, -436, -535, -562, -608]
hist_total_opex = [e + ro + ca for e, ro, ca in zip(hist_employee, hist_road_ops, hist_corp_admin)]

hist_ebitda = [tr + ox for tr, ox in zip(hist_total_rev, hist_total_opex)]
hist_da = [-856, -880, -905, -930, -958]
hist_ebit = [eb + da for eb, da in zip(hist_ebitda, hist_da)]

hist_net_finance = [-907, -821, -963, -1_093, -1_150]
hist_pbt = [eb + nf for eb, nf in zip(hist_ebit, hist_net_finance)]
hist_tax = [7, -141, -270, -249, -282]
hist_npat = [pb + tx for pb, tx in zip(hist_pbt, hist_tax)]

# Row map for IS  (starting at row 5)
IS_START = 5

is_items = [
    ("Revenue", None, None, True, False),
    ("Toll revenue", "A$m", hist_toll_rev, False, False),
    ("Other revenue", "A$m", hist_other_rev, False, False),
    ("Total Revenue", "A$m", hist_total_rev, False, True),
    ("", None, None, False, False),
    ("Operating Expenses", None, None, True, False),
    ("Employee costs", "A$m", hist_employee, False, False),
    ("Road operating costs", "A$m", hist_road_ops, False, False),
    ("Corporate & admin costs", "A$m", hist_corp_admin, False, False),
    ("Total Operating Expenses", "A$m", hist_total_opex, False, True),
    ("", None, None, False, False),
    ("EBITDA", "A$m", hist_ebitda, False, True),
    ("Depreciation & amortisation", "A$m", hist_da, False, False),
    ("EBIT", "A$m", hist_ebit, False, True),
    ("", None, None, False, False),
    ("Net finance costs", "A$m", hist_net_finance, False, False),
    ("Profit / (Loss) before tax", "A$m", hist_pbt, False, True),
    ("Income tax (expense) / benefit", "A$m", hist_tax, False, False),
    ("Net Profit / (Loss) After Tax", "A$m", hist_npat, False, True),
]

# Absolute row numbers  (for cross-referencing)
IS_ROW = {}  # name -> absolute row

r = IS_START
for label, unit, hist, is_section, is_total in is_items:
    ws_is.cell(row=r, column=LABEL_COL, value=label)
    if unit:
        ws_is.cell(row=r, column=UNIT_COL, value=unit)
    IS_ROW[label] = r

    if is_section:
        style_section_row(ws_is, r, MAX_COL)
    elif is_total:
        for c in range(1, MAX_COL + 1):
            ws_is.cell(row=r, column=c).font = total_font
            ws_is.cell(row=r, column=c).border = thick_bottom

    # Write historical values
    if hist is not None:
        for i, v in enumerate(hist):
            col = HIST_START + i
            ws_is.cell(row=r, column=col, value=v).number_format = acct_fmt

    r += 1

# Pre-compute BS and CF row positions from label lists.
# This avoids hard-coding row numbers for cross-sheet references.
# The labels must match exactly those used when building each sheet.
_bs_labels = [
    "ASSETS", "Current Assets", "Cash & cash equivalents",
    "Trade & other receivables", "Other current assets", "Total Current Assets",
    "", "Non-Current Assets", "Property, plant & equipment",
    "Intangible assets (concessions)", "Investments in joint ventures",
    "Other non-current assets", "Total Non-Current Assets", "",
    "Total Assets", "", "LIABILITIES", "Current Liabilities",
    "Trade & other payables", "Current borrowings", "Other current liabilities",
    "Total Current Liabilities", "", "Non-Current Liabilities",
    "Non-current borrowings", "Other non-current liabilities",
    "Total Non-Current Liabilities",
    "Total Borrowings (current + non-current)", "", "Total Liabilities",
    "", "EQUITY", "Share capital", "Retained earnings / (losses)",
    "Reserves", "Total Equity", "", "Total Liabilities & Equity",
    "", "Balance Sheet Check (Assets - L&E)",
]
_bs_row_map = {lbl: 5 + i for i, lbl in enumerate(_bs_labels)}
BS_NCA_ROW = _bs_row_map["Total Non-Current Assets"]
BS_DEBT_ROW = _bs_row_map["Total Borrowings (current + non-current)"]

_cf_labels = [
    "OPERATING ACTIVITIES", "Net profit / (loss) after tax", "Add back: D&A",
    "Changes in working capital", "Other operating adjustments",
    "Net Cash from Operating Activities", "",
    "INVESTING ACTIVITIES", "Capital expenditure", "Other investing activities",
    "Net Cash from Investing Activities", "",
    "FINANCING ACTIVITIES", "Proceeds from / (repayment of) borrowings",
    "Dividends / distributions paid", "Equity issuance (DRP & placements)",
    "Net Cash from Financing Activities", "",
    "Net increase / (decrease) in cash", "FX & other adjustments",
    "Opening cash balance", "Closing Cash Balance",
]
_cf_row_map = {lbl: 5 + i for i, lbl in enumerate(_cf_labels)}
CF_CLOSE_CASH_ROW = _cf_row_map["Closing Cash Balance"]

# -- Forecast formulas (FY26-FY30) --
# Toll revenue: prior year * (1 + toll growth assumption)
for fc_idx in range(5):
    col = FC_START + fc_idx
    prev_col_letter = get_column_letter(col - 1)
    assum_col_letter = get_column_letter(col)

    # Toll revenue  (row IS_ROW["Toll revenue"])
    row_tr = IS_ROW["Toll revenue"]
    formula = f"={prev_col_letter}{row_tr}*(1+Assumptions!{assum_col_letter}{AROW_TOLL_GR})"
    ws_is.cell(row=row_tr, column=col, value=formula).number_format = acct_fmt

    # Other revenue
    row_or = IS_ROW["Other revenue"]
    formula = f"={prev_col_letter}{row_or}*(1+Assumptions!{assum_col_letter}{AROW_OTHER_GR})"
    ws_is.cell(row=row_or, column=col, value=formula).number_format = acct_fmt

    # Total Revenue = Toll + Other
    row_trev = IS_ROW["Total Revenue"]
    formula = f"={get_column_letter(col)}{row_tr}+{get_column_letter(col)}{row_or}"
    ws_is.cell(row=row_trev, column=col, value=formula).number_format = acct_fmt

    # Operating expenses: total opex = revenue * opex%
    # We'll compute total opex first, then split into sub-items proportionally
    row_emp = IS_ROW["Employee costs"]
    row_road = IS_ROW["Road operating costs"]
    row_corp = IS_ROW["Corporate & admin costs"]
    row_topex = IS_ROW["Total Operating Expenses"]
    cl = get_column_letter(col)

    # Employee costs = revenue * employee cost %
    formula = f"=-{cl}{row_trev}*Assumptions!{cl}{AROW_OPEX_PCT}*Assumptions!{cl}{AROW_EMP_SHARE}"
    ws_is.cell(row=row_emp, column=col, value=formula).number_format = acct_fmt

    # Road operating costs ≈ 33% of total opex
    formula = f"=-{cl}{row_trev}*Assumptions!{cl}{AROW_OPEX_PCT}*Assumptions!{cl}{AROW_ROAD_SHARE}"
    ws_is.cell(row=row_road, column=col, value=formula).number_format = acct_fmt

    # Corporate & admin = remainder  (opex% * rev) - employee - road ops
    formula = f"=-{cl}{row_trev}*Assumptions!{cl}{AROW_OPEX_PCT}-{cl}{row_emp}-{cl}{row_road}"
    ws_is.cell(row=row_corp, column=col, value=formula).number_format = acct_fmt

    # Total operating expenses
    formula = f"={cl}{row_emp}+{cl}{row_road}+{cl}{row_corp}"
    ws_is.cell(row=row_topex, column=col, value=formula).number_format = acct_fmt

    # EBITDA
    row_ebitda = IS_ROW["EBITDA"]
    formula = f"={cl}{row_trev}+{cl}{row_topex}"
    ws_is.cell(row=row_ebitda, column=col, value=formula).number_format = acct_fmt

    # D&A  (linked to BS non-current assets via assumption %)
    # D&A = -(opening NCA * D&A %)  -- opening NCA is prior year's closing NCA
    row_da = IS_ROW["Depreciation & amortisation"]
    prev_cl = get_column_letter(col - 1)
    formula = f"=-'Balance Sheet'!{prev_cl}{BS_NCA_ROW}*Assumptions!{cl}{AROW_DA_PCT}"
    ws_is.cell(row=row_da, column=col, value=formula).number_format = acct_fmt

    # EBIT
    row_ebit = IS_ROW["EBIT"]
    formula = f"={cl}{row_ebitda}+{cl}{row_da}"
    ws_is.cell(row=row_ebit, column=col, value=formula).number_format = acct_fmt

    # Net finance costs = -(avg debt * cost of debt)
    # Use opening debt to avoid circular reference
    row_nfc = IS_ROW["Net finance costs"]
    formula = f"=-'Balance Sheet'!{prev_cl}{BS_DEBT_ROW}*Assumptions!{cl}{AROW_COD}"
    ws_is.cell(row=row_nfc, column=col, value=formula).number_format = acct_fmt

    # PBT
    row_pbt = IS_ROW["Profit / (Loss) before tax"]
    formula = f"={cl}{row_ebit}+{cl}{row_nfc}"
    ws_is.cell(row=row_pbt, column=col, value=formula).number_format = acct_fmt

    # Tax
    row_tax = IS_ROW["Income tax (expense) / benefit"]
    formula = f"=-{cl}{row_pbt}*Assumptions!{cl}{AROW_TAX}"
    ws_is.cell(row=row_tax, column=col, value=formula).number_format = acct_fmt

    # NPAT
    row_npat = IS_ROW["Net Profit / (Loss) After Tax"]
    formula = f"={cl}{row_pbt}+{cl}{row_tax}"
    ws_is.cell(row=row_npat, column=col, value=formula).number_format = acct_fmt

    # Shade forecast columns
    for rr in range(IS_START, r):
        mark_forecast_cols(ws_is, rr, col, col)

set_col_widths(ws_is, {"A": 36, "B": 10})
for i in range(HIST_START, MAX_COL + 1):
    ws_is.column_dimensions[get_column_letter(i)].width = 15

# ---------------------------------------------------------------------------
# 3.  BALANCE SHEET  (Sheet 3)
# ---------------------------------------------------------------------------
ws_bs = wb.create_sheet("Balance Sheet")
ws_bs.sheet_properties.tabColor = "548235"

ws_bs.merge_cells("A1:L1")
ws_bs["A1"] = "Transurban Group – Balance Sheet"
ws_bs["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws_bs.row_dimensions[1].height = 30

ws_bs.merge_cells("A2:L2")
ws_bs["A2"] = "A$ millions  |  As at 30 June"
ws_bs["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

r = 4
ws_bs.cell(row=r, column=LABEL_COL, value="Balance Sheet")
for i, fy in enumerate(FY_LABELS):
    ws_bs.cell(row=r, column=HIST_START + i, value=fy)
style_header_row(ws_bs, r, MAX_COL)

# Historical balance sheet data (A$m, approximate)
hist_cash = [2_548, 3_145, 2_780, 2_350, 2_520]
hist_recv = [213, 224, 260, 269, 286]
hist_other_ca = [195, 210, 225, 230, 240]
hist_total_ca = [c + rc + o for c, rc, o in zip(hist_cash, hist_recv, hist_other_ca)]

hist_ppe = [6_125, 6_850, 7_900, 8_100, 8_050]
hist_intangibles = [22_450, 23_100, 24_200, 24_650, 24_800]
hist_invest_jv = [1_820, 1_750, 1_680, 1_620, 1_580]
hist_other_nca = [1_250, 1_380, 1_520, 1_580, 1_620]
hist_total_nca = [p + ig + j + o for p, ig, j, o in
                  zip(hist_ppe, hist_intangibles, hist_invest_jv, hist_other_nca)]

hist_total_assets = [ca + nca for ca, nca in zip(hist_total_ca, hist_total_nca)]

hist_payables = [485, 510, 555, 530, 550]
hist_current_debt = [1_250, 1_400, 1_600, 1_300, 1_200]
hist_other_cl = [580, 620, 680, 710, 740]
hist_total_cl = [p + d + o for p, d, o in zip(hist_payables, hist_current_debt, hist_other_cl)]

hist_nc_debt = [19_580, 21_230, 23_430, 23_730, 23_830]
hist_total_borrow = [cd + nd for cd, nd in zip(hist_current_debt, hist_nc_debt)]
hist_other_ncl = [2_350, 2_480, 2_620, 2_700, 2_780]
hist_total_ncl = [d + o for d, o in zip(hist_nc_debt, hist_other_ncl)]

hist_total_liab = [cl + ncl for cl, ncl in zip(hist_total_cl, hist_total_ncl)]

hist_share_cap = [13_845, 14_250, 14_680, 15_020, 15_310]
hist_reserves = [358, 380, 380, 400, 412]
# Retained earnings is the balancing plug so that Assets = L&E exactly
hist_retained = [
    ta - tl - sc - rs
    for ta, tl, sc, rs in zip(hist_total_assets, hist_total_liab, hist_share_cap, hist_reserves)
]
hist_total_eq = [s + re + rs for s, re, rs in zip(hist_share_cap, hist_retained, hist_reserves)]

hist_total_le = [tl + te for tl, te in zip(hist_total_liab, hist_total_eq)]

bs_items = [
    ("ASSETS", None, None, True, False),
    ("Current Assets", None, None, True, False),
    ("Cash & cash equivalents", "A$m", hist_cash, False, False),
    ("Trade & other receivables", "A$m", hist_recv, False, False),
    ("Other current assets", "A$m", hist_other_ca, False, False),
    ("Total Current Assets", "A$m", hist_total_ca, False, True),
    ("", None, None, False, False),
    ("Non-Current Assets", None, None, True, False),
    ("Property, plant & equipment", "A$m", hist_ppe, False, False),
    ("Intangible assets (concessions)", "A$m", hist_intangibles, False, False),
    ("Investments in joint ventures", "A$m", hist_invest_jv, False, False),
    ("Other non-current assets", "A$m", hist_other_nca, False, False),
    ("Total Non-Current Assets", "A$m", hist_total_nca, False, True),
    ("", None, None, False, False),
    ("Total Assets", "A$m", hist_total_assets, False, True),
    ("", None, None, False, False),
    ("LIABILITIES", None, None, True, False),
    ("Current Liabilities", None, None, True, False),
    ("Trade & other payables", "A$m", hist_payables, False, False),
    ("Current borrowings", "A$m", hist_current_debt, False, False),
    ("Other current liabilities", "A$m", hist_other_cl, False, False),
    ("Total Current Liabilities", "A$m", hist_total_cl, False, True),
    ("", None, None, False, False),
    ("Non-Current Liabilities", None, None, True, False),
    ("Non-current borrowings", "A$m", hist_nc_debt, False, False),
    ("Other non-current liabilities", "A$m", hist_other_ncl, False, False),
    ("Total Non-Current Liabilities", "A$m", hist_total_ncl, False, True),
    ("Total Borrowings (current + non-current)", "A$m", hist_total_borrow, False, True),
    ("", None, None, False, False),
    ("Total Liabilities", "A$m", hist_total_liab, False, True),
    ("", None, None, False, False),
    ("EQUITY", None, None, True, False),
    ("Share capital", "A$m", hist_share_cap, False, False),
    ("Retained earnings / (losses)", "A$m", hist_retained, False, False),
    ("Reserves", "A$m", hist_reserves, False, False),
    ("Total Equity", "A$m", hist_total_eq, False, True),
    ("", None, None, False, False),
    ("Total Liabilities & Equity", "A$m", hist_total_le, False, True),
    ("", None, None, False, False),
    ("Balance Sheet Check (Assets - L&E)", "A$m", None, False, True),
]

BS_ROW = {}
r = 5
for label, unit, hist, is_section, is_total in bs_items:
    ws_bs.cell(row=r, column=LABEL_COL, value=label)
    if unit:
        ws_bs.cell(row=r, column=UNIT_COL, value=unit)
    BS_ROW[label] = r

    if is_section:
        style_section_row(ws_bs, r, MAX_COL)
    elif is_total:
        for c in range(1, MAX_COL + 1):
            ws_bs.cell(row=r, column=c).font = total_font
            ws_bs.cell(row=r, column=c).border = thick_bottom if label != "Balance Sheet Check (Assets - L&E)" else double_bottom

    if hist is not None:
        for i, v in enumerate(hist):
            ws_bs.cell(row=r, column=HIST_START + i, value=v).number_format = acct_fmt

    r += 1

# Verify pre-computed row positions match actual
actual_bs_nca_row = BS_ROW["Total Non-Current Assets"]
actual_bs_debt_row = BS_ROW["Total Borrowings (current + non-current)"]
assert actual_bs_nca_row == BS_NCA_ROW, f"BS NCA row mismatch: {actual_bs_nca_row} != {BS_NCA_ROW}"
assert actual_bs_debt_row == BS_DEBT_ROW, f"BS debt row mismatch: {actual_bs_debt_row} != {BS_DEBT_ROW}"

# Now write BS check formulas for historical
row_ta = BS_ROW["Total Assets"]
row_tle = BS_ROW["Total Liabilities & Equity"]
row_check = BS_ROW["Balance Sheet Check (Assets - L&E)"]
for col in range(HIST_START, HIST_END + 1):
    cl = get_column_letter(col)
    ws_bs.cell(row=row_check, column=col,
               value=f"={cl}{row_ta}-{cl}{row_tle}").number_format = acct_fmt

# ---- FORECAST FORMULAS FOR BALANCE SHEET (FY26-FY30) ----
row_cash = BS_ROW["Cash & cash equivalents"]
row_recv = BS_ROW["Trade & other receivables"]
row_oca = BS_ROW["Other current assets"]
row_tca = BS_ROW["Total Current Assets"]
row_ppe = BS_ROW["Property, plant & equipment"]
row_intang = BS_ROW["Intangible assets (concessions)"]
row_jv = BS_ROW["Investments in joint ventures"]
row_onca = BS_ROW["Other non-current assets"]
row_tnca = BS_ROW["Total Non-Current Assets"]
row_ta = BS_ROW["Total Assets"]
row_pay = BS_ROW["Trade & other payables"]
row_cd = BS_ROW["Current borrowings"]
row_ocl = BS_ROW["Other current liabilities"]
row_tcl = BS_ROW["Total Current Liabilities"]
row_ncd = BS_ROW["Non-current borrowings"]
row_oncl = BS_ROW["Other non-current liabilities"]
row_tncl = BS_ROW["Total Non-Current Liabilities"]
row_tb = BS_ROW["Total Borrowings (current + non-current)"]
row_tl = BS_ROW["Total Liabilities"]
row_sc = BS_ROW["Share capital"]
row_re = BS_ROW["Retained earnings / (losses)"]
row_rsv = BS_ROW["Reserves"]
row_teq = BS_ROW["Total Equity"]
row_tle = BS_ROW["Total Liabilities & Equity"]

# IS rows needed
IS_ROW_NPAT = IS_ROW["Net Profit / (Loss) After Tax"]
IS_ROW_DA = IS_ROW["Depreciation & amortisation"]
IS_ROW_TREV = IS_ROW["Total Revenue"]

for fc_idx in range(5):
    col = FC_START + fc_idx
    cl = get_column_letter(col)
    prev_cl = get_column_letter(col - 1)

    # -- ASSETS --
    # Cash comes from Cash Flow Statement (closing cash)
    ws_bs.cell(row=row_cash, column=col,
               value=f"='Cash Flow Statement'!{cl}{CF_CLOSE_CASH_ROW}").number_format = acct_fmt

    # Trade receivables = (Revenue / 365) * receivable days
    ws_bs.cell(row=row_recv, column=col,
               value=f"='Income Statement'!{cl}{IS_ROW_TREV}/365*Assumptions!{cl}{AROW_REC_DAYS}").number_format = acct_fmt

    # Other current assets: grow at 3% p.a.
    ws_bs.cell(row=row_oca, column=col,
               value=f"={prev_cl}{row_oca}*(1+Assumptions!{cl}{AROW_OCA_GR})").number_format = acct_fmt

    # Total CA
    ws_bs.cell(row=row_tca, column=col,
               value=f"={cl}{row_cash}+{cl}{row_recv}+{cl}{row_oca}").number_format = acct_fmt

    # PP&E = prior PP&E + capex + D&A (D&A is negative, so adding reduces)
    ws_bs.cell(row=row_ppe, column=col,
               value=f"={prev_cl}{row_ppe}+Assumptions!{cl}{AROW_CAPEX}+'Income Statement'!{cl}{IS_ROW_DA}*Assumptions!{cl}{AROW_DA_PPE}").number_format = acct_fmt

    # Intangibles = prior - amortisation (60% of D&A allocated to intangibles)
    ws_bs.cell(row=row_intang, column=col,
               value=f"={prev_cl}{row_intang}+'Income Statement'!{cl}{IS_ROW_DA}*Assumptions!{cl}{AROW_DA_INTANG}").number_format = acct_fmt

    # JV investments: stable, slight decline
    ws_bs.cell(row=row_jv, column=col,
               value=f"={prev_cl}{row_jv}*(1+Assumptions!{cl}{AROW_JV_GR})").number_format = acct_fmt

    # Other NCA: grow at 2%
    ws_bs.cell(row=row_onca, column=col,
               value=f"={prev_cl}{row_onca}*(1+Assumptions!{cl}{AROW_ONCA_GR})").number_format = acct_fmt

    # Total NCA
    ws_bs.cell(row=row_tnca, column=col,
               value=f"={cl}{row_ppe}+{cl}{row_intang}+{cl}{row_jv}+{cl}{row_onca}").number_format = acct_fmt

    # Total Assets
    ws_bs.cell(row=row_ta, column=col,
               value=f"={cl}{row_tca}+{cl}{row_tnca}").number_format = acct_fmt

    # -- LIABILITIES --
    # Trade payables = (Total opex / 365) * payable days
    IS_ROW_TOPEX = IS_ROW["Total Operating Expenses"]
    ws_bs.cell(row=row_pay, column=col,
               value=f"=-'Income Statement'!{cl}{IS_ROW_TOPEX}/365*Assumptions!{cl}{AROW_PAY_DAYS}").number_format = acct_fmt

    # Current borrowings: assume stable proportion (~5% of total debt)
    ws_bs.cell(row=row_cd, column=col,
               value=f"=({prev_cl}{row_tb}+Assumptions!{cl}{AROW_NET_DEBT})*Assumptions!{cl}{AROW_CURR_BORROW_PCT}").number_format = acct_fmt

    # Other current liabilities: grow at 3%
    ws_bs.cell(row=row_ocl, column=col,
               value=f"={prev_cl}{row_ocl}*(1+Assumptions!{cl}{AROW_OCL_GR})").number_format = acct_fmt

    # Total CL
    ws_bs.cell(row=row_tcl, column=col,
               value=f"={cl}{row_pay}+{cl}{row_cd}+{cl}{row_ocl}").number_format = acct_fmt

    # Non-current borrowings = prior total debt + net debt issuance - current borrowings
    ws_bs.cell(row=row_ncd, column=col,
               value=f"={prev_cl}{row_tb}+Assumptions!{cl}{AROW_NET_DEBT}-{cl}{row_cd}").number_format = acct_fmt

    # Other NCL: grow at 2%
    ws_bs.cell(row=row_oncl, column=col,
               value=f"={prev_cl}{row_oncl}*(1+Assumptions!{cl}{AROW_ONCL_GR})").number_format = acct_fmt

    # Total NCL
    ws_bs.cell(row=row_tncl, column=col,
               value=f"={cl}{row_ncd}+{cl}{row_oncl}").number_format = acct_fmt

    # Total Borrowings
    ws_bs.cell(row=row_tb, column=col,
               value=f"={cl}{row_cd}+{cl}{row_ncd}").number_format = acct_fmt

    # Total Liabilities
    ws_bs.cell(row=row_tl, column=col,
               value=f"={cl}{row_tcl}+{cl}{row_tncl}").number_format = acct_fmt

    # -- EQUITY --
    # Share capital: prior + assumed equity raise (DRP ~1% dilution)
    ws_bs.cell(row=row_sc, column=col,
               value=f"={prev_cl}{row_sc}*(1+Assumptions!{cl}{AROW_DRP_RATE})").number_format = acct_fmt

    # Retained earnings = prior RE + NPAT - dividends paid
    # Dividends = DPS * shares / 100 (DPS in cents)
    ws_bs.cell(row=row_re, column=col,
               value=f"={prev_cl}{row_re}+'Income Statement'!{cl}{IS_ROW_NPAT}-Assumptions!{cl}{AROW_DPS}*Assumptions!{cl}{AROW_SHARES}/100").number_format = acct_fmt

    # Reserves: stable
    ws_bs.cell(row=row_rsv, column=col,
               value=f"={prev_cl}{row_rsv}").number_format = acct_fmt

    # Total Equity
    ws_bs.cell(row=row_teq, column=col,
               value=f"={cl}{row_sc}+{cl}{row_re}+{cl}{row_rsv}").number_format = acct_fmt

    # Total L&E
    ws_bs.cell(row=row_tle, column=col,
               value=f"={cl}{row_tl}+{cl}{row_teq}").number_format = acct_fmt

    # BS Check
    ws_bs.cell(row=row_check, column=col,
               value=f"={cl}{row_ta}-{cl}{row_tle}").number_format = acct_fmt

    # Shade
    for rr in range(5, r):
        mark_forecast_cols(ws_bs, rr, col, col)

set_col_widths(ws_bs, {"A": 40, "B": 10})
for i in range(HIST_START, MAX_COL + 1):
    ws_bs.column_dimensions[get_column_letter(i)].width = 15

# ---------------------------------------------------------------------------
# 4.  CASH FLOW STATEMENT  (Sheet 4)
# ---------------------------------------------------------------------------
ws_cf = wb.create_sheet("Cash Flow Statement")
ws_cf.sheet_properties.tabColor = "BF8F00"

ws_cf.merge_cells("A1:L1")
ws_cf["A1"] = "Transurban Group – Cash Flow Statement"
ws_cf["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws_cf.row_dimensions[1].height = 30

ws_cf.merge_cells("A2:L2")
ws_cf["A2"] = "A$ millions  |  Fiscal year ends 30 June"
ws_cf["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

r = 4
ws_cf.cell(row=r, column=LABEL_COL, value="Cash Flow Statement")
for i, fy in enumerate(FY_LABELS):
    ws_cf.cell(row=r, column=HIST_START + i, value=fy)
style_header_row(ws_cf, r, MAX_COL)

# Historical cash flow data
hist_cfo_npat = hist_npat  # starting point
hist_cfo_da = [-x for x in hist_da]  # add back (positive)
hist_wc_change = [-45, 30, -60, 25, -15]
hist_other_ops = [120, 135, 150, 160, 170]
hist_net_cfo = [np + da + wc + oth for np, da, wc, oth in
                zip(hist_cfo_npat, hist_cfo_da, hist_wc_change, hist_other_ops)]

hist_capex = [-628, -1_092, -1_805, -1_420, -1_200]
hist_other_inv = [-150, -80, -120, -100, -90]
hist_net_cfi = [cx + oi for cx, oi in zip(hist_capex, hist_other_inv)]

hist_debt_proc = [1_200, 2_050, 2_800, 600, 300]
hist_div_paid = [-792, -1_032, -1_218, -1_276, -1_313]
hist_equity_iss = [405, 450, 430, 340, 290]
hist_net_cff = [dp + dv + eq for dp, dv, eq in
                zip(hist_debt_proc, hist_div_paid, hist_equity_iss)]

hist_net_change = [o + i + f for o, i, f in zip(hist_net_cfo, hist_net_cfi, hist_net_cff)]
# Opening cash = prior year closing cash on BS
hist_open_cash = [2_285]  # FY21 opening (FY20 closing cash)
for k in range(4):
    hist_open_cash.append(hist_cash[k])  # closing cash of prior year = opening of next
# FX & other = closing cash (BS) - opening cash - net change
hist_fx_other = [c - oc - nc for c, oc, nc in zip(hist_cash, hist_open_cash, hist_net_change)]

cf_items = [
    ("OPERATING ACTIVITIES", None, None, True, False),
    ("Net profit / (loss) after tax", "A$m", hist_cfo_npat, False, False),
    ("Add back: D&A", "A$m", hist_cfo_da, False, False),
    ("Changes in working capital", "A$m", hist_wc_change, False, False),
    ("Other operating adjustments", "A$m", hist_other_ops, False, False),
    ("Net Cash from Operating Activities", "A$m", hist_net_cfo, False, True),
    ("", None, None, False, False),
    ("INVESTING ACTIVITIES", None, None, True, False),
    ("Capital expenditure", "A$m", hist_capex, False, False),
    ("Other investing activities", "A$m", hist_other_inv, False, False),
    ("Net Cash from Investing Activities", "A$m", hist_net_cfi, False, True),
    ("", None, None, False, False),
    ("FINANCING ACTIVITIES", None, None, True, False),
    ("Proceeds from / (repayment of) borrowings", "A$m", hist_debt_proc, False, False),
    ("Dividends / distributions paid", "A$m", hist_div_paid, False, False),
    ("Equity issuance (DRP & placements)", "A$m", hist_equity_iss, False, False),
    ("Net Cash from Financing Activities", "A$m", hist_net_cff, False, True),
    ("", None, None, False, False),
    ("Net increase / (decrease) in cash", "A$m", hist_net_change, False, True),
    ("FX & other adjustments", "A$m", hist_fx_other, False, False),
    ("Opening cash balance", "A$m", hist_open_cash, False, False),
    ("Closing Cash Balance", "A$m", hist_cash, False, True),
]

CF_ROW = {}
r = 5
for label, unit, hist, is_section, is_total in cf_items:
    ws_cf.cell(row=r, column=LABEL_COL, value=label)
    if unit:
        ws_cf.cell(row=r, column=UNIT_COL, value=unit)
    CF_ROW[label] = r

    if is_section:
        style_section_row(ws_cf, r, MAX_COL)
    elif is_total:
        for c in range(1, MAX_COL + 1):
            ws_cf.cell(row=r, column=c).font = total_font
            ws_cf.cell(row=r, column=c).border = thick_bottom if label != "Closing Cash Balance" else double_bottom

    if hist is not None:
        for i, v in enumerate(hist):
            ws_cf.cell(row=r, column=HIST_START + i, value=v).number_format = acct_fmt

    r += 1

# Verify pre-computed CF row positions match actual
actual_cf_close_row = CF_ROW["Closing Cash Balance"]
assert actual_cf_close_row == CF_CLOSE_CASH_ROW, \
    f"CF close row mismatch: {actual_cf_close_row} != {CF_CLOSE_CASH_ROW}"

# ---- FORECAST FORMULAS FOR CASH FLOW (FY26-FY30) ----
cf_npat_row = CF_ROW["Net profit / (loss) after tax"]
cf_da_row = CF_ROW["Add back: D&A"]
cf_wc_row = CF_ROW["Changes in working capital"]
cf_other_ops_row = CF_ROW["Other operating adjustments"]
cf_net_ops_row = CF_ROW["Net Cash from Operating Activities"]
cf_capex_row = CF_ROW["Capital expenditure"]
cf_other_inv_row = CF_ROW["Other investing activities"]
cf_net_inv_row = CF_ROW["Net Cash from Investing Activities"]
cf_debt_row = CF_ROW["Proceeds from / (repayment of) borrowings"]
cf_div_row = CF_ROW["Dividends / distributions paid"]
cf_equity_row = CF_ROW["Equity issuance (DRP & placements)"]
cf_net_fin_row = CF_ROW["Net Cash from Financing Activities"]
cf_net_change_row = CF_ROW["Net increase / (decrease) in cash"]
cf_fx_row = CF_ROW["FX & other adjustments"]
cf_open_row = CF_ROW["Opening cash balance"]
cf_close_row = CF_ROW["Closing Cash Balance"]

for fc_idx in range(5):
    col = FC_START + fc_idx
    cl = get_column_letter(col)
    prev_cl = get_column_letter(col - 1)

    # -- OPERATING --
    # NPAT from IS
    ws_cf.cell(row=cf_npat_row, column=col,
               value=f"='Income Statement'!{cl}{IS_ROW_NPAT}").number_format = acct_fmt

    # D&A add-back (positive) = negative of IS D&A
    ws_cf.cell(row=cf_da_row, column=col,
               value=f"=-'Income Statement'!{cl}{IS_ROW_DA}").number_format = acct_fmt

    # Working capital change = -(change in receivables) + (change in payables)
    ws_cf.cell(row=cf_wc_row, column=col,
               value=f"=-('Balance Sheet'!{cl}{row_recv}-'Balance Sheet'!{prev_cl}{row_recv})+('Balance Sheet'!{cl}{row_pay}-'Balance Sheet'!{prev_cl}{row_pay})").number_format = acct_fmt

    # Other operating adjustments: held stable
    ws_cf.cell(row=cf_other_ops_row, column=col,
               value=f"={prev_cl}{cf_other_ops_row}*(1+Assumptions!{cl}{AROW_OTHER_OPS_GR})").number_format = acct_fmt

    # Net CFO
    ws_cf.cell(row=cf_net_ops_row, column=col,
               value=f"={cl}{cf_npat_row}+{cl}{cf_da_row}+{cl}{cf_wc_row}+{cl}{cf_other_ops_row}").number_format = acct_fmt

    # -- INVESTING --
    # Capex from assumptions (negative)
    ws_cf.cell(row=cf_capex_row, column=col,
               value=f"=-Assumptions!{cl}{AROW_CAPEX}").number_format = acct_fmt

    # Other investing: held roughly stable
    ws_cf.cell(row=cf_other_inv_row, column=col,
               value=f"={prev_cl}{cf_other_inv_row}").number_format = acct_fmt

    # Net CFI
    ws_cf.cell(row=cf_net_inv_row, column=col,
               value=f"={cl}{cf_capex_row}+{cl}{cf_other_inv_row}").number_format = acct_fmt

    # -- FINANCING --
    # Debt proceeds/repayment from assumptions
    ws_cf.cell(row=cf_debt_row, column=col,
               value=f"=Assumptions!{cl}{AROW_NET_DEBT}").number_format = acct_fmt

    # Dividends paid = -(DPS * shares / 100)
    ws_cf.cell(row=cf_div_row, column=col,
               value=f"=-Assumptions!{cl}{AROW_DPS}*Assumptions!{cl}{AROW_SHARES}/100").number_format = acct_fmt

    # Equity issuance ≈ prior share capital * 1% DRP
    ws_cf.cell(row=cf_equity_row, column=col,
               value=f"='Balance Sheet'!{prev_cl}{row_sc}*Assumptions!{cl}{AROW_EQUITY_ISS}").number_format = acct_fmt

    # Net CFF
    ws_cf.cell(row=cf_net_fin_row, column=col,
               value=f"={cl}{cf_debt_row}+{cl}{cf_div_row}+{cl}{cf_equity_row}").number_format = acct_fmt

    # Net change in cash
    ws_cf.cell(row=cf_net_change_row, column=col,
               value=f"={cl}{cf_net_ops_row}+{cl}{cf_net_inv_row}+{cl}{cf_net_fin_row}").number_format = acct_fmt

    # FX / other: assume nil in forecast
    ws_cf.cell(row=cf_fx_row, column=col, value=0).number_format = acct_fmt

    # Opening cash = prior period closing cash on BS
    ws_cf.cell(row=cf_open_row, column=col,
               value=f"='Balance Sheet'!{prev_cl}{row_cash}").number_format = acct_fmt

    # Closing cash
    ws_cf.cell(row=cf_close_row, column=col,
               value=f"={cl}{cf_open_row}+{cl}{cf_net_change_row}+{cl}{cf_fx_row}").number_format = acct_fmt

    # Shade
    for rr in range(5, r):
        mark_forecast_cols(ws_cf, rr, col, col)

set_col_widths(ws_cf, {"A": 40, "B": 10})
for i in range(HIST_START, MAX_COL + 1):
    ws_cf.column_dimensions[get_column_letter(i)].width = 15

# ---------------------------------------------------------------------------
# 5.  NOTES TO THE FINANCIAL STATEMENTS  (Sheet 5)
# ---------------------------------------------------------------------------
ws_notes = wb.create_sheet("Notes")
ws_notes.sheet_properties.tabColor = "7030A0"  # Purple

ws_notes.merge_cells("A1:L1")
ws_notes["A1"] = "Transurban Group – Notes to the Financial Statements"
ws_notes["A1"].font = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
ws_notes.row_dimensions[1].height = 30

ws_notes.merge_cells("A2:L2")
ws_notes["A2"] = "A$ millions  |  Fiscal year ends 30 June"
ws_notes["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")

# Header row (row 4)
r = 4
ws_notes.cell(row=r, column=LABEL_COL, value="Notes")
ws_notes.cell(row=r, column=UNIT_COL, value="")
for i, fy in enumerate(FY_LABELS):
    ws_notes.cell(row=r, column=HIST_START + i, value=fy)
style_header_row(ws_notes, r, MAX_COL)

# Historical construction revenue data
hist_construction_rev = [180, 320, 540, 420, 350]

# Historical segment data
hist_melb_rev = [820, 920, 1080, 1150, 1220]
hist_melb_ebitda = [620, 710, 840, 900, 960]
hist_syd_rev = [1050, 1220, 1520, 1680, 1800]
hist_syd_ebitda = [750, 890, 1120, 1250, 1350]
hist_bris_rev = [430, 480, 560, 600, 640]
hist_bris_ebitda = [310, 350, 410, 440, 470]
hist_na_rev = [478, 521, 627, 652, 690]
hist_na_ebitda = [120, 150, 205, 230, 260]

# Historical intangible assets data
hist_intang_opening = [22100, 22630, 23310, 24460, 25340]
hist_intang_additions = [530, 680, 1150, 880, 720]

# Historical borrowings maturity data
hist_maturity_1_2yr = [1100, 1250, 1400, 1150, 1050]
hist_maturity_2_5yr = [5200, 5650, 6300, 6400, 6500]

# Historical borrowing costs data
hist_cap_borrow_costs = [45, 78, 125, 95, 80]

# Historical tax adjustment data
hist_non_ded_amort = [215, 210, 205, 195, 190]
hist_tax_concessions = [-50, -48, -45, -42, -40]
hist_other_perm_diff = [25, 20, 15, 10, 8]

# Historical commitments data
hist_cap_commit = [2800, 3200, 2500, 1800, 1500]
hist_op_lease_commit = [85, 90, 95, 100, 105]
hist_contingent_liab = [150, 150, 160, 160, 170]

# Define all note items with their data
notes_items = [
    # Note 1: Revenue Breakdown
    ("REVENUE BREAKDOWN", None, None, True, False, False),
    ("Toll revenue", "A$m", "link_is_toll", False, False, False),
    ("Construction revenue", "A$m", hist_construction_rev, False, False, False),
    ("Other revenue", "A$m", "link_is_other", False, False, False),
    ("Total Revenue", "A$m", "link_is_total_rev", False, True, False),
    ("", None, None, False, False, False),
    
    # Note 2: Segment Reporting
    ("SEGMENT REPORTING", None, None, True, False, False),
    ("Melbourne (CityLink)", None, None, False, False, False),
    ("  Revenue", "A$m", hist_melb_rev, False, False, False),
    ("  EBITDA", "A$m", hist_melb_ebitda, False, False, False),
    ("Sydney", None, None, False, False, False),
    ("  Revenue", "A$m", hist_syd_rev, False, False, False),
    ("  EBITDA", "A$m", hist_syd_ebitda, False, False, False),
    ("Brisbane", None, None, False, False, False),
    ("  Revenue", "A$m", hist_bris_rev, False, False, False),
    ("  EBITDA", "A$m", hist_bris_ebitda, False, False, False),
    ("North America", None, None, False, False, False),
    ("  Revenue", "A$m", hist_na_rev, False, False, False),
    ("  EBITDA", "A$m", hist_na_ebitda, False, False, False),
    ("Total segment revenue", "A$m", "sum_segment_rev", False, True, False),
    ("Total segment EBITDA", "A$m", "sum_segment_ebitda", False, True, False),
    ("Reconciliation to IS EBITDA", "A$m", "link_is_ebitda", False, True, False),
    ("", None, None, False, False, False),
    
    # Note 3: Intangible Assets
    ("INTANGIBLE ASSETS (CONCESSION RIGHTS)", None, None, True, False, False),
    ("Opening balance", "A$m", hist_intang_opening, False, False, False),
    ("Additions (capitalised construction)", "A$m", hist_intang_additions, False, False, False),
    ("Amortisation charge", "A$m", "link_is_da_60pct", False, False, False),
    ("Closing balance", "A$m", "calc_intang_close", False, True, False),
    ("Cross-check to BS", "A$m", "link_bs_intang", False, True, True),
    ("", None, None, False, False, False),
    
    # Note 4: Borrowings
    ("BORROWINGS", None, None, True, False, False),
    ("Current borrowings", "A$m", "link_bs_current_debt", False, False, False),
    ("Non-current borrowings", "A$m", "link_bs_nc_debt", False, False, False),
    ("Total borrowings", "A$m", "link_bs_total_debt", False, True, False),
    ("", None, None, False, False, False),
    ("Maturity Profile", None, None, False, False, False),
    ("  Within 1 year", "A$m", "link_current_debt", False, False, False),
    ("  1-2 years", "A$m", hist_maturity_1_2yr, False, False, False),
    ("  2-5 years", "A$m", hist_maturity_2_5yr, False, False, False),
    ("  Over 5 years", "A$m", "calc_over_5yr", False, False, False),
    ("  Total (maturity check)", "A$m", "sum_maturity", False, True, True),
    ("", None, None, False, False, False),
    ("Borrowing Costs", None, None, False, False, False),
    ("  Interest expense", "A$m", "link_is_interest", False, False, False),
    ("  Capitalised borrowing costs", "A$m", hist_cap_borrow_costs, False, False, False),
    ("  Effective interest rate", "%", "link_assum_cod", False, False, False),
    ("", None, None, False, False, False),
    
    # Note 5: Income Tax
    ("INCOME TAX", None, None, True, False, False),
    ("Profit before tax", "A$m", "link_is_pbt", False, False, False),
    ("Tax at statutory rate (30%)", "A$m", "calc_tax_30pct", False, False, False),
    ("Tax effect adjustments:", None, None, False, False, False),
    ("  Non-deductible amortisation", "A$m", hist_non_ded_amort, False, False, False),
    ("  Tax concessions & offsets", "A$m", hist_tax_concessions, False, False, False),
    ("  Other permanent differences", "A$m", hist_other_perm_diff, False, False, False),
    ("Total tax adjustments", "A$m", "sum_tax_adj", False, True, False),
    ("Income tax expense", "A$m", "link_is_tax", False, True, False),
    ("Effective tax rate", "%", "calc_etr", False, False, False),
    ("ETR per Assumptions", "%", "link_assum_tax", False, False, True),
    ("", None, None, False, False, False),
    
    # Note 6: Dividends/Distributions
    ("DIVIDENDS / DISTRIBUTIONS", None, None, True, False, False),
    ("DPS (cents per security)", "A¢", "link_assum_dps", False, False, False),
    ("Securities on issue (m)", "m", "link_assum_shares", False, False, False),
    ("Total distributions paid", "A$m", "link_cf_div", False, False, False),
    ("Payout ratio (% of NPAT)", "%", "calc_payout", False, False, False),
    ("Franking credits", "A$m", [0, 0, 0, 0, 0], False, False, False),
    ("", None, None, False, False, False),
    
    # Note 7: Commitments & Contingencies
    ("COMMITMENTS & CONTINGENCIES", None, None, True, False, False),
    ("Capital commitments", "A$m", hist_cap_commit, False, False, False),
    ("Operating lease commitments", "A$m", hist_op_lease_commit, False, False, False),
    ("Contingent liabilities", "A$m", hist_contingent_liab, False, False, False),
]

NOTES_ROW = {}
r = 5
for label, unit, data, is_section, is_total, is_check in notes_items:
    ws_notes.cell(row=r, column=LABEL_COL, value=label)
    if unit:
        ws_notes.cell(row=r, column=UNIT_COL, value=unit)
    NOTES_ROW[label] = r
    
    if is_section:
        style_section_row(ws_notes, r, MAX_COL)
    elif is_total:
        for c in range(1, MAX_COL + 1):
            ws_notes.cell(row=r, column=c).font = total_font
            if is_check:
                ws_notes.cell(row=r, column=c).border = double_bottom
            else:
                ws_notes.cell(row=r, column=c).border = thick_bottom
    
    # Write historical values or formulas
    if data is not None and isinstance(data, list):
        # Hard-coded historical data
        for i, v in enumerate(data):
            col = HIST_START + i
            cell = ws_notes.cell(row=r, column=col, value=v)
            if unit == "%":
                cell.number_format = pct_fmt
            elif unit in ["A$m", "A¢", "m", "days"]:
                cell.number_format = acct_fmt if unit == "A$m" else num_fmt
    
    r += 1

# Now add formulas for historical and forecast periods
# Note 1: Revenue Breakdown
row_toll = NOTES_ROW["Toll revenue"]
row_construction = NOTES_ROW["Construction revenue"]
row_other_rev = NOTES_ROW["Other revenue"]
row_total_rev_note = NOTES_ROW["Total Revenue"]

for col_idx in range(10):  # FY21-FY30
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    prev_cl = get_column_letter(col - 1)
    
    # Toll revenue - link to IS
    ws_notes.cell(row=row_toll, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['Toll revenue']}").number_format = acct_fmt
    
    # Construction revenue - historical hard-coded, forecast grows at 3%
    if col >= FC_START:
        ws_notes.cell(row=row_construction, column=col,
                      value=f"={prev_cl}{row_construction}*(1+Assumptions!{cl}{AROW_CONSTR_GR})").number_format = acct_fmt
    
    # Other revenue - link to IS
    ws_notes.cell(row=row_other_rev, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['Other revenue']}").number_format = acct_fmt
    
    # Total Revenue - link to IS
    ws_notes.cell(row=row_total_rev_note, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['Total Revenue']}").number_format = acct_fmt

# Note 2: Segment Reporting (historical only, no forecast formulas)
# Calculate segment row positions based on the known structure
row_melb_rev = NOTES_ROW["Melbourne (CityLink)"] + 1  # Melbourne revenue is one row after the label
row_melb_ebitda = row_melb_rev + 1
row_syd_rev = NOTES_ROW["Sydney"] + 1  # Sydney revenue is one row after the label
row_syd_ebitda = row_syd_rev + 1
row_bris_rev = NOTES_ROW["Brisbane"] + 1  # Brisbane revenue is one row after the label
row_bris_ebitda = row_bris_rev + 1
row_na_rev = NOTES_ROW["North America"] + 1  # North America revenue is one row after the label
row_na_ebitda = row_na_rev + 1
row_total_seg_rev = NOTES_ROW["Total segment revenue"]
row_total_seg_ebitda = NOTES_ROW["Total segment EBITDA"]
row_recon_ebitda = NOTES_ROW["Reconciliation to IS EBITDA"]

for col_idx in range(5):  # Historical only FY21-FY25
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    
    # Total segment revenue = sum of all segment revenues
    ws_notes.cell(row=row_total_seg_rev, column=col,
                  value=f"={cl}{row_melb_rev}+{cl}{row_syd_rev}+{cl}{row_bris_rev}+{cl}{row_na_rev}").number_format = acct_fmt
    
    # Total segment EBITDA = sum of all segment EBITDAs
    ws_notes.cell(row=row_total_seg_ebitda, column=col,
                  value=f"={cl}{row_melb_ebitda}+{cl}{row_syd_ebitda}+{cl}{row_bris_ebitda}+{cl}{row_na_ebitda}").number_format = acct_fmt
    
    # Reconciliation to IS EBITDA
    ws_notes.cell(row=row_recon_ebitda, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['EBITDA']}").number_format = acct_fmt

# Note 3: Intangible Assets
row_intang_open = NOTES_ROW["Opening balance"]
row_intang_add = NOTES_ROW["Additions (capitalised construction)"]
row_intang_amort = NOTES_ROW["Amortisation charge"]
row_intang_close = NOTES_ROW["Closing balance"]
row_intang_check = NOTES_ROW["Cross-check to BS"]

for col_idx in range(10):  # FY21-FY30
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    prev_cl = get_column_letter(col - 1)
    
    # Opening balance - for FY21 use hard-coded, for others use prior year closing
    if col > HIST_START:
        ws_notes.cell(row=row_intang_open, column=col,
                      value=f"={prev_cl}{row_intang_close}").number_format = acct_fmt
    
    # Additions - historical hard-coded, forecast links to construction revenue
    if col >= FC_START:
        ws_notes.cell(row=row_intang_add, column=col,
                      value=f"={cl}{row_construction}").number_format = acct_fmt
    
    # Amortisation charge - link to IS D&A * 0.60
    ws_notes.cell(row=row_intang_amort, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['Depreciation & amortisation']}*Assumptions!{cl}{AROW_DA_INTANG}").number_format = acct_fmt
    
    # Closing balance = Opening + Additions + Amortisation (amort is negative)
    ws_notes.cell(row=row_intang_close, column=col,
                  value=f"={cl}{row_intang_open}+{cl}{row_intang_add}+{cl}{row_intang_amort}").number_format = acct_fmt
    
    # Cross-check to BS
    ws_notes.cell(row=row_intang_check, column=col,
                  value=f"='Balance Sheet'!{cl}{BS_ROW['Intangible assets (concessions)']}").number_format = acct_fmt

# Note 4: Borrowings
row_curr_debt = NOTES_ROW["Current borrowings"]
row_nc_debt = NOTES_ROW["Non-current borrowings"]
row_total_debt = NOTES_ROW["Total borrowings"]
row_mat_within_1 = NOTES_ROW["  Within 1 year"]
row_mat_1_2 = NOTES_ROW["  1-2 years"]
row_mat_2_5 = NOTES_ROW["  2-5 years"]
row_mat_over_5 = NOTES_ROW["  Over 5 years"]
row_mat_total = NOTES_ROW["  Total (maturity check)"]
row_interest = NOTES_ROW["  Interest expense"]
row_cap_costs = NOTES_ROW["  Capitalised borrowing costs"]
row_eff_rate = NOTES_ROW["  Effective interest rate"]

for col_idx in range(10):  # FY21-FY30
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    
    # Link to BS borrowings
    ws_notes.cell(row=row_curr_debt, column=col,
                  value=f"='Balance Sheet'!{cl}{BS_ROW['Current borrowings']}").number_format = acct_fmt
    ws_notes.cell(row=row_nc_debt, column=col,
                  value=f"='Balance Sheet'!{cl}{BS_ROW['Non-current borrowings']}").number_format = acct_fmt
    ws_notes.cell(row=row_total_debt, column=col,
                  value=f"='Balance Sheet'!{cl}{BS_ROW['Total Borrowings (current + non-current)']}").number_format = acct_fmt
    
    # Maturity profile
    # Within 1 year = current borrowings
    ws_notes.cell(row=row_mat_within_1, column=col,
                  value=f"={cl}{row_curr_debt}").number_format = acct_fmt
    
    # 1-2 years - historical hard-coded, forecast = 5% of total
    if col >= FC_START:
        ws_notes.cell(row=row_mat_1_2, column=col,
                      value=f"={cl}{row_total_debt}*Assumptions!{cl}{AROW_MAT_1_2}").number_format = acct_fmt
    
    # 2-5 years - historical hard-coded, forecast = 26% of total
    if col >= FC_START:
        ws_notes.cell(row=row_mat_2_5, column=col,
                      value=f"={cl}{row_total_debt}*Assumptions!{cl}{AROW_MAT_2_5}").number_format = acct_fmt
    
    # Over 5 years = Total - within 1yr - 1-2yr - 2-5yr
    ws_notes.cell(row=row_mat_over_5, column=col,
                  value=f"={cl}{row_total_debt}-{cl}{row_mat_within_1}-{cl}{row_mat_1_2}-{cl}{row_mat_2_5}").number_format = acct_fmt
    
    # Total maturity check
    ws_notes.cell(row=row_mat_total, column=col,
                  value=f"={cl}{row_mat_within_1}+{cl}{row_mat_1_2}+{cl}{row_mat_2_5}+{cl}{row_mat_over_5}").number_format = acct_fmt
    
    # Interest expense - link to IS with sign flip
    ws_notes.cell(row=row_interest, column=col,
                  value=f"=-'Income Statement'!{cl}{IS_ROW['Net finance costs']}").number_format = acct_fmt
    
    # Capitalised borrowing costs - historical hard-coded, forecast formula
    if col >= FC_START:
        ws_notes.cell(row=row_cap_costs, column=col,
                      value=f"=Assumptions!{cl}{AROW_CAPEX}*Assumptions!{cl}{AROW_COD}*Assumptions!{cl}{AROW_CAP_BORR}").number_format = acct_fmt
    
    # Effective interest rate - link to Assumptions
    ws_notes.cell(row=row_eff_rate, column=col,
                  value=f"=Assumptions!{cl}{AROW_COD}").number_format = pct_fmt

# Note 5: Income Tax
row_pbt = NOTES_ROW["Profit before tax"]
row_tax_30 = NOTES_ROW["Tax at statutory rate (30%)"]
row_non_ded = NOTES_ROW["  Non-deductible amortisation"]
row_tax_conc = NOTES_ROW["  Tax concessions & offsets"]
row_other_perm = NOTES_ROW["  Other permanent differences"]
row_total_adj = NOTES_ROW["Total tax adjustments"]
row_tax_exp = NOTES_ROW["Income tax expense"]
row_etr = NOTES_ROW["Effective tax rate"]
row_etr_assum = NOTES_ROW["ETR per Assumptions"]

for col_idx in range(10):  # FY21-FY30
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    
    # PBT - link to IS
    ws_notes.cell(row=row_pbt, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['Profit / (Loss) before tax']}").number_format = acct_fmt
    
    # Tax at 30%
    ws_notes.cell(row=row_tax_30, column=col,
                  value=f"={cl}{row_pbt}*-Assumptions!{cl}{AROW_STAT_TAX}").number_format = acct_fmt
    
    # Adjustments - historical hard-coded, forecast holds flat
    if col >= FC_START:
        ws_notes.cell(row=row_non_ded, column=col,
                      value=f"=Assumptions!{cl}{AROW_NON_DED}").number_format = acct_fmt
        ws_notes.cell(row=row_tax_conc, column=col,
                      value=f"=Assumptions!{cl}{AROW_TAX_CONC}").number_format = acct_fmt
        ws_notes.cell(row=row_other_perm, column=col,
                      value=f"=Assumptions!{cl}{AROW_OTHER_PERM}").number_format = acct_fmt
    
    # Total tax adjustments
    ws_notes.cell(row=row_total_adj, column=col,
                  value=f"={cl}{row_non_ded}+{cl}{row_tax_conc}+{cl}{row_other_perm}").number_format = acct_fmt
    
    # Income tax expense - link to IS
    ws_notes.cell(row=row_tax_exp, column=col,
                  value=f"='Income Statement'!{cl}{IS_ROW['Income tax (expense) / benefit']}").number_format = acct_fmt
    
    # Effective tax rate
    ws_notes.cell(row=row_etr, column=col,
                  value=f"={cl}{row_tax_exp}/{cl}{row_pbt}").number_format = pct_fmt
    
    # ETR per Assumptions
    ws_notes.cell(row=row_etr_assum, column=col,
                  value=f"=Assumptions!{cl}{AROW_TAX}").number_format = pct_fmt

# Note 6: Dividends/Distributions
row_dps = NOTES_ROW["DPS (cents per security)"]
row_shares = NOTES_ROW["Securities on issue (m)"]
row_dist_paid = NOTES_ROW["Total distributions paid"]
row_payout = NOTES_ROW["Payout ratio (% of NPAT)"]
row_franking = NOTES_ROW["Franking credits"]

for col_idx in range(10):  # FY21-FY30
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    
    # DPS - link to Assumptions
    ws_notes.cell(row=row_dps, column=col,
                  value=f"=Assumptions!{cl}{AROW_DPS}").number_format = num_fmt
    
    # Securities on issue - link to Assumptions
    ws_notes.cell(row=row_shares, column=col,
                  value=f"=Assumptions!{cl}{AROW_SHARES}").number_format = num_fmt
    
    # Total distributions paid - link to CF
    ws_notes.cell(row=row_dist_paid, column=col,
                  value=f"='Cash Flow Statement'!{cl}{CF_ROW['Dividends / distributions paid']}").number_format = acct_fmt
    
    # Payout ratio
    ws_notes.cell(row=row_payout, column=col,
                  value=f"=-{cl}{row_dist_paid}/'Income Statement'!{cl}{IS_ROW['Net Profit / (Loss) After Tax']}").number_format = pct_fmt
    
    # Franking credits - hold at 0 for forecast
    if col >= FC_START:
        ws_notes.cell(row=row_franking, column=col, value=0).number_format = acct_fmt

# Note 7: Commitments & Contingencies
row_cap_commit = NOTES_ROW["Capital commitments"]
row_op_lease = NOTES_ROW["Operating lease commitments"]
row_contingent = NOTES_ROW["Contingent liabilities"]

for col_idx in range(10):  # FY21-FY30
    col = HIST_START + col_idx
    cl = get_column_letter(col)
    prev_cl = get_column_letter(col - 1)
    
    # Capital commitments - historical hard-coded, forecast links to capex
    if col >= FC_START:
        ws_notes.cell(row=row_cap_commit, column=col,
                      value=f"=Assumptions!{cl}{AROW_CAPEX}*Assumptions!{cl}{AROW_CAP_COMMIT}").number_format = acct_fmt
    
    # Operating lease commitments - historical hard-coded, forecast grows at 3%
    if col >= FC_START:
        ws_notes.cell(row=row_op_lease, column=col,
                      value=f"={prev_cl}{row_op_lease}*(1+Assumptions!{cl}{AROW_OP_LEASE_GR})").number_format = acct_fmt
    
    # Contingent liabilities - historical hard-coded, forecast holds flat
    if col >= FC_START:
        ws_notes.cell(row=row_contingent, column=col,
                      value=f"=Assumptions!{cl}{AROW_CONTINGENT}").number_format = acct_fmt

# Shade forecast columns for all note rows
for rr in range(5, r):
    for col in range(FC_START, FC_END + 1):
        mark_forecast_cols(ws_notes, rr, col, col)

# Set column widths
set_col_widths(ws_notes, {"A": 40, "B": 10})
for i in range(HIST_START, MAX_COL + 1):
    ws_notes.column_dimensions[get_column_letter(i)].width = 15

# ---------------------------------------------------------------------------
# 6.  FINAL FORMATTING PASS
# ---------------------------------------------------------------------------

# Freeze panes and print settings for each sheet
for ws in [ws_is, ws_bs, ws_cf, ws_notes]:
    ws.freeze_panes = "C5"
    ws.sheet_view.showGridLines = False
# Assumptions sheet has headers at row 12, so freeze below row 12
ws_a.freeze_panes = "C13"
ws_a.sheet_view.showGridLines = False

# ---------------------------------------------------------------------------
# 7.  SAVE
# ---------------------------------------------------------------------------
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "Transurban_Group_3Way_Financial_Model.xlsx")
wb.save(OUTPUT_PATH)
print(f"Model saved to: {OUTPUT_PATH}")
print("Sheets:", [s.title for s in wb.worksheets])
